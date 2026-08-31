"""AUTO-GENERATED single-file serverless extractor. DO NOT EDIT BY HAND.
Regenerate with:  python3 api/build_bundle.py
Sources: api/_core/*.py"""
import hashlib
import io
import json
import os
import re
import sys
import tempfile
import traceback
from datetime import datetime

_IMPORT_ERRS: list[str] = []


# ============================ supabase client =========================
"""Minimal Supabase REST client (stdlib only): auth, PostgREST, storage."""
import json
import urllib.error
import urllib.request


class SupabaseError(Exception):
    pass


class SB:
    def __init__(self, url: str, service_key: str, timeout: int = 30):
        self.url = url.rstrip("/")
        self.key = service_key
        self.timeout = timeout

    def _req(self, method, path, body=None, headers=None, raw=None, content_type=None, expect_json=True):
        hdrs = {"apikey": self.key, "Authorization": f"Bearer {self.key}"}
        if headers:
            hdrs.update(headers)
        data = None
        if raw is not None:
            data = raw
            hdrs["Content-Type"] = content_type or "application/octet-stream"
        elif body is not None:
            data = json.dumps(body).encode()
            hdrs["Content-Type"] = "application/json"
        req = urllib.request.Request(f"{self.url}{path}", data=data, method=method, headers=hdrs)
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                payload = resp.read()
                if not expect_json:
                    return payload
                return json.loads(payload) if payload else None
        except urllib.error.HTTPError as e:
            detail = e.read().decode(errors="replace")[:400]
            raise SupabaseError(f"{method} {path} -> {e.code}: {detail}") from e

    # ---- auth ----
    def get_user(self, access_token: str):
        """Returns user dict or None."""
        try:
            return self._req("GET", "/auth/v1/user",
                             headers={"Authorization": f"Bearer {access_token}"})
        except SupabaseError:
            return None

    # ---- PostgREST ----
    def select(self, table: str, query: str, token: str | None = None):
        hdrs = {}
        if token:
            hdrs["Authorization"] = f"Bearer {token}"
        rows = self._req("GET", f"/rest/v1/{table}?{query}", headers=hdrs)
        return rows or []

    def insert(self, table: str, rows: list, upsert: bool = False, returning: str = "representation"):
        hdrs = {"Prefer": f"return={returning}" + (",resolution=merge-duplicates" if upsert else "")}
        return self._req("POST", f"/rest/v1/{table}", body=rows, headers=hdrs)

    def rpc(self, fn: str, args: dict):
        return self._req("POST", f"/rest/v1/rpc/{fn}", body=args)

    # ---- storage ----
    def storage_upload(self, bucket: str, path: str, data: bytes, content_type: str):
        return self._req("POST", f"/storage/v1/object/{bucket}/{path}",
                         raw=data, content_type=content_type, expect_json=True)

    def create_signed_url(self, bucket: str, path: str, expires_in: int = 600):
        r = self._req("POST", f"/storage/v1/object/sign/{bucket}/{path}",
                      body={"expiresIn": expires_in})
        return r.get("signedURL") or r.get("signedUrl")


# ============================ xlsx builder ============================
"""Schema-compliant XLSX builder (openpyxl only - no pandas)."""
import io

try:
    from openpyxl import Workbook
except Exception as _e:
    Workbook = None
    _IMPORT_ERRS.append(f"openpyxl: {_e}")
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TX_COLS = [
    ("trans_date", "Trans Date", 12),
    ("ref_number", "Ref Number", 14),
    ("raw_details", "Original Details", 52),
    ("transaction_category", "Transaction Category", 28),
    ("party_name", "Counterparty / Person", 30),
    ("reference_id", "System Ref", 22),
    ("value_date", "Value Date", 12),
    ("withdrawal_dr", "Withdrawal (DR)", 15),
    ("deposit_cr", "Deposit (CR)", 15),
    ("balance", "Balance", 15),
]


def build_xlsx(payload) -> bytes:
    """payload: {bank_name, account_summary, transactions} -> xlsx bytes."""
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)

    ws = wb.active
    ws.title = "Account Summary"
    s = payload.get("account_summary", {})
    rows = [("Bank", payload.get("bank_name", ""))] + list(s.items())
    ws.append(["Field", "Value"])
    for k, v in rows:
        ws.append([k, v if v is not None else ""])
    for c in ws[1]:
        c.fill, c.font = hdr_fill, hdr_font
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 48
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Transactions")
    ws2.append([h for _, h, _ in TX_COLS])
    for c in ws2[1]:
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(vertical="center")
    for t in payload.get("transactions", []):
        ws2.append([t.get(k, "") for k, _, _ in TX_COLS])
    money_idx = [i + 1 for i, (k, _, _) in enumerate(TX_COLS)
                 if k in ("withdrawal_dr", "deposit_cr", "balance")]
    for i, (k, _, w) in enumerate(TX_COLS, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        if k in ("withdrawal_dr", "deposit_cr", "balance"):
            for r in range(2, ws2.max_row + 1):
                ws2.cell(row=r, column=i).number_format = "#,##0.00"
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================ multi-bank parsers ======================
"""Multi-bank statement parsers: GTBank, ADB, OmniBSIC, Ecobank, MTN MoMo.
Ported 1:1 from the verified pipeline; process() returns payload instead of writing files."""
import json
import os
import re

try:
    import pymupdf
except Exception as _e:  # guarded: reported via /api/extract health
    pymupdf = None
    _IMPORT_ERRS.append(f"pymupdf: {_e}")


"""
Multi-bank statement extractor for AVE MARIA SCHOOL statements.
Banks: GTBank (GT), ADB, OmniBSIC, Ecobank, MTN MoMo  (+ FBN handled by parse_statement.py)

Design: shared word/line engine with per-bank column configs + row anchors,
description attachment by nearest anchor, running-balance verification.
"""
import re
from datetime import datetime

try:
    import pymupdf
except Exception as _e:  # guarded: reported via /api/extract health
    pymupdf = None
    _IMPORT_ERRS.append(f"pymupdf: {_e}")

MONTHS = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}
IDX_MONTHS = {i + 1: m for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}


def fmt_date_ddmmyyyy(s):
    """01-04-2026 -> 01-Apr-2026 (day-first); DD-Mon-YYYY passes through."""
    if re.match(r"^\d{2}-[A-Za-z]{3}-\d{4}$", s):
        return s
    d, m, y = s.split("-")
    return f"{d}-{IDX_MONTHS[int(m)]}-{y}"


def fmt_date_slashmon(s):
    """01/JUN/2026 -> 01-Jun-2026."""
    d, m, y = s.split("/")
    return f"{d}-{m.capitalize()}-{y}"


def fmt_date_iso(s):
    """2024-01-04 -> 04-Jan-2024."""
    y, m, d = s.split("-")
    return f"{d}-{IDX_MONTHS[int(m)]}-{y}"


def num(s):
    """'1,234.56' / '(2.00)' / '2017.6' / '' -> float"""
    s = (s or "").strip().replace(",", "").replace(" ", "")
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    if not s or s in {"-", "."}:
        return 0.0
    try:
        v = float(s)
    except ValueError:
        return 0.0
    return -v if neg else v


def group_lines(words, tol=2.0):
    """words: [(x0,y0,x1,y1,text)] -> list of (y, [words sorted by x])"""
    lines = {}
    for w in words:
        placed = False
        for k in list(lines):
            if abs(k - w[1]) < tol:
                lines[k].append(w)
                placed = True
                break
        if not placed:
            lines[w[1]] = [w]
    return [(y, sorted(ws, key=lambda w: w[0])) for y, ws in sorted(lines.items())]


def bucket(ws, edges):
    """assign words to columns by x0 with right-edge correction for numbers"""
    cols = [[] for _ in edges[:-1]]
    for w in ws:
        idx = len(cols) - 1
        for i in range(len(cols) - 1):
            if w[0] < edges[i + 1] - 2:
                idx = i
                break
        cols[idx].append(w)
    return [" ".join(w[4] for w in c) for c in cols]


NUM_RE = re.compile(r"^-?\(?[\d,]+\.\d{2}\)?$|^-?\(?[\d,]+\)?$")


class ColumnBank:
    """Generic engine: fixed column edges, anchor predicate, nearest-anchor desc attach."""
    edges = None
    date_col = 0
    balance_col = -1
    amount_cols = ()
    desc_cols = ()
    anchor_date_re = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$|^\d{2}-\d{2}-\d{4}$")
    skip_lines = ()
    max_gap = 40.0
    attach_mode = "nearest"

    def line_texts(self, ws):
        return bucket(ws, self.edges)

    def is_anchor(self, cols):
        return bool(self.anchor_date_re.match(cols[self.date_col])) and \
            NUM_RE.match(cols[self.balance_col].replace(" ", "")) is not None

    def is_junk(self, y, cols):
        joined = " ".join(c for c in cols if c).upper()
        return any(pat in joined for pat in self.skip_lines)

    def parse(self, doc):
        """Two-pass: collect lines, then attach desc lines to nearest anchor (either side)."""
        all_lines = []  # (pno, y, cols)
        for pno, page in enumerate(doc):
            for y, ws in group_lines(page.get_text("words")):
                cols = self.line_texts(ws)
                if self.is_junk(y, cols):
                    continue
                all_lines.append((pno, y, cols))
        rows = []
        anchors = []  # (pno, y, row)
        for pno, y, cols in all_lines:
            if self.is_anchor(cols):
                row = self.make_row(pno, y, cols)
                rows.append(row)
                anchors.append([pno, y, row])
        # attach desc-only lines
        last_above = None
        for pno, y, cols in all_lines:
            if self.is_anchor(cols):
                last_above = next((a for a in reversed(anchors) if a[0] == pno and a[1] <= y), last_above)
                continue
            if not self.has_desc(cols):
                continue
            if self.attach_mode == "above":
                if last_above is not None and last_above[0] == pno:
                    self.absorb(y, last_above[2], cols)
                continue
            best, bd = None, None
            for a in anchors:
                if a[0] != pno:
                    continue
                d = abs(a[1] - y)
                if bd is None or d < bd:
                    best, bd = a, d
            if best is not None and bd <= self.max_gap:
                self.absorb(y, best[2], cols)
        return rows

    def has_desc(self, cols):
        return bool(" ".join(cols[i] for i in self.desc_cols).strip())

    def make_row(self, pno, y, cols):
        raise NotImplementedError

    def absorb(self, y, row, cols, anchors_y):
        desc = " ".join(cols[i] for i in self.desc_cols).strip()
        if not desc:
            return
        row.setdefault("_extra", []).append(desc)

    def finalize(self, rows):
        """attach _extra lines to nearest anchor (within max_gap)"""
        for i, r in enumerate(rows):
            extras = []
            for y, desc, pno in r.get("_extra_meta", []):
                extras.append(desc)
            base = r.get("raw_details", "")
            pieces = [p for p in [base] + extras if p]
            r["raw_details"] = re.sub(r"\s+", " ", " ".join(pieces)).strip()
        return rows


# ---------------------------------------------------------------- GTBank ----
class GTB(ColumnBank):
    attach_mode = "above"
    edges = [0, 100, 200, 285, 365, 450, 10000]  # date value debit credit balance remarks
    date_col, balance_col = 0, 4
    amount_cols = (2, 3)
    desc_cols = (5,)
    anchor_date_re = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")
    skip_lines = ("TOTAL:", "Statement Period", "Trans. Date")

    def make_row(self, pno, y, cols):
        return {"page": pno, "y": y, "_pre": [], "_post": [], "_extra_meta": [],
                "trans": fmt_date_ddmmyyyy(cols[0]) if "-" in cols[0] else cols[0],
                "value": fmt_date_ddmmyyyy(cols[1]) if "-" in cols[1] else cols[1],
                "dr_raw": cols[2].strip(), "cr_raw": cols[3].strip(),
                "bal": num(cols[4]),
                "raw_details": cols[5].strip()}

    def absorb(self, y, row, cols, anchors=None):
        desc = cols[5].strip()
        if desc:
            row["_post"].append(desc)

    def compose(self, r):
        parts = [re.sub(r"\s+", " ", p).strip() for p in [r["raw_details"]] + r["_post"]]
        return re.sub(r"\s+", " ", " ".join(p for p in parts if p)).strip()

    def finalize(self, rows):
        for r in rows:
            r["raw_details"] = self.compose(r)
            r["withdrawal_dr"] = round(num(r["dr_raw"]), 2)
            r["deposit_cr"] = round(num(r["cr_raw"]), 2)
            r["balance"] = round(r["bal"], 2)
            r["value_date"] = r["value"]
        return rows


# ------------------------------------------------------------------ ADB ----
class ADB(ColumnBank):
    edges = [0, 55, 90, 246, 336, 404, 462, 520, 10000]
    max_gap = 14.0
    # date branch desc ref value debit credit balance
    date_col, balance_col = 0, 7
    amount_cols = (5, 6)
    desc_cols = (2, 3)
    anchor_date_re = re.compile(r"^\d{2}-\d{2}-\d{4}$")
    skip_lines = ("STATEMENT OF ACCOUNT", "Period From", "Account No", "NOT FOR VISA")

    def make_row(self, pno, y, cols):
        return {"page": pno, "y": y, "_pre": [], "_post": [], "_extra_meta": [],
                "trans": cols[0].strip(), "branch": cols[1].strip(),
                "raw_details": cols[2].strip(), "ref": cols[3].strip(),
                "value": cols[4].strip(), "dr_raw": cols[5].strip(),
                "cr_raw": cols[6].strip(), "bal": num(cols[7])}

    def absorb(self, y, row, cols, anchors=None):
        desc_cols_txt = " ".join(cols[i] for i in self.desc_cols).strip()
        if not desc_cols_txt:
            return
        (row["_pre"] if y < row["y"] else row["_post"]).append(desc_cols_txt)

    def compose(self, r):
        parts = [re.sub(r"\s+", " ", p).strip() for p in r["_pre"] + [r["raw_details"]] + r["_post"]]
        return re.sub(r"\s+", " ", " ".join(p for p in parts if p)).strip()

    def finalize(self, rows):
        for r in rows:
            r["raw_details"] = self.compose(r)
            r["trans"] = fmt_date_ddmmyyyy(r["trans"])
            r["value_date"] = fmt_date_ddmmyyyy(r["value"]) if re.match(r"^\d{2}-\d{2}-\d{4}$", r["value"]) else r["trans"]
            r["withdrawal_dr"] = round(num(r["dr_raw"]), 2)
            r["deposit_cr"] = round(num(r["cr_raw"]), 2)
            r["balance"] = round(r["bal"], 2)
            r["ref_number"] = r["ref"]
        return rows


# ------------------------------------------------------------- OmniBSIC ----
class OMNI(ColumnBank):
    edges = [0, 75, 238, 300, 368, 440, 10000]
    max_gap = 10.0
    # booking desc value DR CR bal
    date_col, balance_col = 0, 5
    amount_cols = (3, 4)
    desc_cols = (1,)
    anchor_date_re = re.compile(r"^\d{2}-\d{2}-\d{4}$")
    skip_lines = ("ACCOUNT STATEMENT", "Name Of Customer", "Balance Brought Forward",
                  "Booking Date", "INTERNAL USE ONLY")
    watermark = {"INTERNAL", "USE", "ONLY", "USE ONLY", "INTERNAL USE"}

    def make_row(self, pno, y, cols):
        return {"page": pno, "y": y, "_pre": [], "_post": [], "_extra_meta": [],
                "trans": cols[0].strip(), "raw_details": cols[1].strip(),
                "value": cols[2].strip(), "dr_raw": cols[3].strip(),
                "cr_raw": cols[4].strip(), "bal": num(cols[5])}

    def absorb(self, y, row, cols, anchors=None):
        desc = cols[1].strip()
        if not desc or desc in self.watermark:
            return
        (row["_pre"] if y < row["y"] else row["_post"]).append(desc)

    def compose(self, r):
        parts = [re.sub(r"\s+", " ", p).strip() for p in r["_pre"] + [r["raw_details"]] + r["_post"]]
        return re.sub(r"\s+", " ", " ".join(p for p in parts if p)).strip()

    def finalize(self, rows):
        for r in rows:
            r["raw_details"] = self.compose(r)
            r["trans"] = fmt_date_ddmmyyyy(r["trans"])
            r["value_date"] = fmt_date_ddmmyyyy(r["value"]) if re.match(r"^\d{2}-\d{2}-\d{4}$", r["value"]) else r["trans"]
            r["withdrawal_dr"] = round(num(r["dr_raw"]), 2)
            r["deposit_cr"] = round(num(r["cr_raw"]), 2)
            r["balance"] = round(r["bal"], 2)
        return rows


# ------------------------------------------------------------- Ecobank ----
ECO_DATE = re.compile(r"^\d{2}/[A-Z]{3}/\d{4}$")
ECO_REF = re.compile(r"\b(H[0-9A-Z]{13,})\b")
ECO_ACCT_FRAG = {"14410", "04968", "722"}


class ECO(ColumnBank):
    edges = [0, 60, 220, 262, 355, 435, 510, 10000]
    max_gap = 30.0
    # date desc inst value debit credit balance
    date_col, balance_col = 0, 6
    amount_cols = (4, 5)
    desc_cols = (1, 2)
    anchor_date_re = ECO_DATE
    skip_lines = ("STATEMENT OF ACCOUNT", "Cust Ac No", "Opening Balance", "Closing Balance",
                  "Cleared Balance", "Uncleared Effects", "Total Withdrawals", "Total Deposits",
                  "Statement Period", "P O BOX", "AVE MARIA SCHOOL LIMITED")

    def make_row(self, pno, y, cols):
        return {"page": pno, "y": y, "_pre": [], "_post": [],
                "trans": fmt_date_slashmon(cols[0].strip()),
                "raw_details": "", "value": "", "dr_raw": cols[4].strip(),
                "cr_raw": cols[5].strip(), "bal": num(cols[6])}

    def absorb(self, y, row, cols, anchors=None):
        desc = cols[1].strip()
        if not desc:
            return
        desc = " ".join(p for p in desc.split() if p not in ECO_ACCT_FRAG)
        if desc:
            (row["_pre"] if y < row["y"] else row["_post"]).append(desc)

    def compose(self, r):
        parts = [re.sub(r"\s+", " ", p).strip() for p in r["_pre"] + [r["raw_details"]] + r["_post"]]
        return re.sub(r"\s+", " ", " ".join(p for p in parts if p)).strip()

    def finalize(self, rows):
        for r in rows:
            r["raw_details"] = self.compose(r)
            r["value_date"] = r["value"] or r["trans"]
            r["withdrawal_dr"] = round(num(r["dr_raw"]), 2)
            r["deposit_cr"] = round(num(r["cr_raw"]), 2)
            r["balance"] = round(r["bal"], 2)
        return rows


# ---------------------------------------------------------- MTN MoMo ----
MOMO_EDGES = [83, 130, 180, 230, 320, 380, 494, 538, 564, 586, 614, 660, 704, 746, 800, 862, 930, 20000]
MOMO_COLS = ["d1", "datetime", "from_acct", "from_name", "from_phone", "type", "amt",
             "from_fee", "to_fee", "tax", "bal_before", "bal_after", "after_acct",
             "after_name", "to_msisdn", "fin_id", "message"]
ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
YYYYMMDD = re.compile(r"^\d{8}$")


def parse_momo(doc):
    rows = []
    for pno, page in enumerate(doc):
        words = page.get_text("words")
        # drop the header words (top of page) for pages > 1 handled by anchor detection
        for y, ws in group_lines(words, tol=2.5):
            cols = bucket(ws, MOMO_EDGES)
            d1 = cols[0].strip()
            if not YYYYMMDD.match(d1):
                # continuation line -> previous row
                if rows and y > 240:
                    extra = " ".join(c for c in cols[3:] if c.strip()).strip()
                    if extra and not ISO_DATE.match(cols[1].strip()):
                        rows[-1]["_cont"].append(extra)
                continue
            rec = dict(zip(MOMO_COLS, [c.strip() for c in cols]))
            rec["page"] = pno
            rec["_cont"] = []
            rows.append(rec)
    out = []
    for r in rows:
        dt = f"{r['d1'][:4]}-{r['d1'][4:6]}-{r['d1'][6:]}"
        trans = fmt_date_iso(dt)
        bal_before, bal_after = num(r["bal_before"]), num(r["bal_after"])
        delta = round(bal_after - bal_before, 2)
        amt = num(r["amt"])
        fee = round(num(r["from_fee"]) + num(r["to_fee"]), 2)
        tax = num(r["tax"])
        dr, cr = (0.0, 0.0)
        if delta >= 0:
            cr = delta
        else:
            dr = -delta
        cont = " ".join(r["_cont"])
        from_side = r["from_name"] or r["from_phone"]
        raw = (f"{r['type']} | From: {from_side} ({r['from_phone']}) Acct {r['from_acct']} | "
               f"To: {r['after_name']} ({r['to_msisdn']}) Acct {r['after_acct']} | "
               f"Bal {bal_before:,.2f} -> {bal_after:,.2f} | Fee {fee:,.2f} Tax {tax:,.2f} | "
               f"FinID {r['fin_id']} Msg {r['message']} {cont}")
        raw = re.sub(r"\s+", " ", raw).strip()
        out.append({
            "page": r["page"], "trans": trans, "time": r["datetime"],
            "raw_details": raw, "withdrawal_dr": round(dr, 2), "deposit_cr": round(cr, 2),
            "balance": round(bal_after, 2), "bal_before": round(bal_before, 2), "value_date": trans,
            "type": r["type"], "amount": amt, "from_name": r["from_name"],
            "from_phone": r["from_phone"], "to_msisdn": r["to_msisdn"],
            "after_name": r["after_name"], "ref": r["from_acct"],
            "delta_match": abs(abs(delta) - amt) < 0.005,
            "fee": fee, "tax": tax,
        })
    return out


# ------------------------------------------------ metadata extraction ----
def between(text, start_label, end_labels):
    i = text.find(start_label)
    if i < 0:
        return ""
    rest = text[i + len(start_label):]
    for el in end_labels:
        j = rest.find(el)
        if j >= 0:
            rest = rest[:j]
    return rest.strip(" :\n").strip()


def meta_adb(doc):
    t = doc[0].get_text()
    addr = between(t, "Customer Address", ["Account Title", "Date\nBranch"])
    addr = re.sub(r"\s+", " ", addr)
    m = re.search(r"Period From :\s*(\d{2}-\d{2}-\d{4})\s*To :\s*(\d{2}-\d{2}-\d{4})", t)
    return {
        "account_name": between(t, "Customer Name", ["Customer Address", "Account Title"]),
        "customer_address": addr,
        "account_number": between(t, "Account No", ["Product Name"]),
        "account_type": "Product " + between(t, "Product Name", ["Currency Name"]) + " - " +
                        between(t, "Branch Name", ["Customer ID"]),
        "currency": between(t, "Currency Name", ["Branch Code"]),
        "statement_period_start": fmt_date_ddmmyyyy(m.group(1)) if m else "",
        "statement_period_end": fmt_date_ddmmyyyy(m.group(2)) if m else "",
        "available_balance": None, "total_credit": None, "total_debit": None,
        "opening_balance": None, "closing_balance": None,
    }


def meta_gt(doc):
    t = doc[0].get_text()
    m = re.search(r"Statement Period:\s*(\d{2}-[A-Za-z]{3}-\d{4})\s*to\s*(\d{2}-[A-Za-z]{3}-\d{4})", t)
    return {
        "account_name": between(t, "CUSTOMER STATEMENT", ["Print. Date"]),
        "customer_address": between(t, "Address", ["Account Type"]),
        "account_number": between(t, "Account No.", ["Address", "Account Type"]),
        "account_type": between(t, "Account Type", ["Currency"]),
        "currency": between(t, "Currency", ["Opening Balance"]),
        "statement_period_start": m.group(1) if m else "",
        "statement_period_end": m.group(2) if m else "",
        "opening_balance": num(between(t, "Opening Balance", ["Closing Balance", "Trans. Date"])),
        "closing_balance": num(between(t, "Closing Balance", ["Trans. Date"])),
        "available_balance": None, "total_credit": None, "total_debit": None,
    }


def meta_omni(doc):
    t = doc[0].get_text()
    m = re.search(r"Start Date\s*:\s*([A-Za-z]+ \d{1,2}, \d{4})", t)
    m2 = re.search(r"End Date\s*:\s*([A-Za-z]+ \d{1,2}, \d{4})", t)
    f = lambda s: datetime.strptime(s, "%b %d, %Y").strftime("%d-%b-%Y") if s else ""
    return {
        "account_name": between(t, "Name Of Customer:", ["Branch Name"]),
        "customer_address": "",
        "account_number": between(t, "Account Number", ["Uncleared Balance"]),
        "account_type": "ACCOUNT STATEMENT - " + between(t, "Branch Name:", ["Available Balance"]).title(),
        "currency": between(t, "Ccy", ["Balance Brought Forward", "Start Date"]),
        "statement_period_start": f(m.group(1)) if m else "",
        "statement_period_end": f(m2.group(1)) if m2 else "",
        "opening_balance": num(between(t, "Balance Brought Forward:", ["Start Date", "Closing Balance"])),
        "closing_balance": num(between(t, "Closing Balance", ["End Date"])),
        "available_balance": num(between(t, "Available Balance", ["Account Number"])),
        "total_credit": None, "total_debit": None,
    }


def meta_eco(doc):
    t = doc[0].get_text()
    m = re.search(r"Statement Period\s*\n?\s*(\d{2}-[A-Z]{3}-\d{4})\s*To\s*(\d{2}-[A-Z]{3}-\d{4})", t)
    return {
        "account_name": "AVE MARIA SCHOOL LIMITED",
        "customer_address": between(t, "AVE MARIA SCHOOL LIMITED", ["STATEMENT OF ACCOUNT", "1"]),
        "account_number": between(t, "Cust Ac No", ["Ccy"]),
        "account_type": "CURRENT ACCOUNT",
        "currency": between(t, "Ccy", ["Opening Balance"]),
        "statement_period_start": fmt_date_slashmon(m.group(1).replace("-", "/")) if m else "",
        "statement_period_end": fmt_date_slashmon(m.group(2).replace("-", "/")) if m else "",
        "opening_balance": num(between(t, "Opening Balance", ["Closing Balance"])),
        "closing_balance": num(between(t, "Closing Balance", ["Cleared Balance"])),
        "available_balance": num(between(t, "Cleared Balance", ["Uncleared Effects"])),
        "total_debit": num(between(t, "Total Withdrawals", ["Total Deposits"])),
        "total_credit": num(between(t, "Total Deposits", ["Statement Period"])),
    }


def meta_momo(doc):
    p = doc[0]
    words = p.get_text("words")
    def near(label, xlo, xhi, pat):
        ys = [w[1] for w in words if w[4] == label]
        if not ys:
            return ""
        y0 = ys[0]
        for w in words:
            if w[1] > y0 and xlo <= w[0] <= xhi and re.match(pat, w[4]):
                return w[4]
        return ""
    m_from = near("From_date", 760, 800, r"\d{1,2}/\d{1,2}/\d{4}")
    m_to = near("To_Date", 808, 850, r"\d{1,2}/\d{1,2}/\d{4}")
    num = near("Customer_Number", 480, 545, r"\d{9,15}")
    label_y = next((w[1] for w in words if w[4] == "profile"), None)
    prof = [w[4] for w in sorted(words, key=lambda w: (w[1], w[0]))
            if 718 <= w[0] < 770 and label_y is not None and label_y + 2 < w[1] < label_y + 45]
    f = lambda s: datetime.strptime(s, "%m/%d/%Y").strftime("%d-%b-%Y") if s else ""
    return {
        "account_name": "AVE MARIA SCHOOL LIMITED (MTN MoMo Merchant)",
        "customer_address": "",
        "account_number": num,
        "account_type": "MTN Mobile Money - " + " ".join(prof),
        "currency": "GHS",
        "statement_period_start": f(m_from),
        "statement_period_end": f(m_to),
        "opening_balance": None, "closing_balance": None,
        "available_balance": None, "total_credit": None, "total_debit": None,
    }


# ------------------------------------------------------- categorization ----
def categorize_gtb(raw, dr=0.0, cr=0.0):
    d = raw.upper()
    if "CASH DEP" in d:
        return "Cash Deposit (3rd Party)"
    if "GIP_" in d or "_GIP" in d:
        return "GhIPSS Credit (Fees)"
    if "RTGSGH" in d:
        return "Interbank Transfer (RTGS)"
    if "E-PRODUCT BUNDLE" in d or "E-BUNDLE" in d:
        return "E-Product Bundle Charge"
    if "GT INWARD CLEARING" in d:
        return "Inward Clearing (Cheque)"
    if "OUT CHQ" in d or "CHQ" in d and dr > 0:
        return "Cheque Payment (Outward)"
    if "COT:" in d:
        return "COT Charge"
    if "STATEMENT CHARGE" in d:
        return "Statement Charge"
    if "WITHDRAWAL" in d or "CASH W/D" in d:
        return "Cash Withdrawal"
    if "TRANSFER" in d or "TRF" in d:
        return "Funds Transfer"
    if cr > 0:
        return "Counter Credit (Deposit)"
    if dr > 0:
        return "Other Debit"
    return "Other"


def categorize_adb(raw, dr=0.0, cr=0.0):
    d = raw.upper()
    if "CASH DEPOSIT BY" in d:
        return "Cash Deposit (3rd Party)"
    if "CHEQUE WITHDRAWAL BY" in d:
        return "Cheque Withdrawal (Self)"
    if "CREDIT INTEREST" in d:
        return "Credit Interest"
    if "T-BILL" in d:
        return "T-Bill Transaction"
    if "COMMISSION ON NRT" in d:
        return "Interbank Transfer Charge (NRT)"
    if "SALARY" in d:
        return "Salary Payment (FTM)"
    if "MOMO" in d:
        return "MoMo Transfer"
    if "ACH NRT" in d or "ACH CREDIT" in d:
        return "ACH/Interbank Credit"
    if "NEGOTIATED BENEFITS TRUST" in d:
        return "Transfer (Negotiated Benefits Trust)"
    if re.search(r"109101001112890\d\s*-\s*TO\s*-", d):
        return "Internal Transfer (Own ADB Accounts)"
    if "MTN-PULL" in d:
        return "MoMo Pull Debit"
    if "M_APP" in d and "INTERNAL" in d:
        return "Mobile App Transfer"
    if "GIP" in d and "INCOMING" in d:
        return "GhIPSS Incoming Credit"
    if "INWARD CLEARING CHEQUE" in d:
        return "Inward Clearing (Cheque Paid)"
    if "ADB CHQ NO" in d:
        return "Cheque Payment (ADB)"
    if "FUNDS TRANSFER FROM" in d or "FUND TRANSFER FROM" in d:
        return "Funds Transfer (Credit)"
    if "ACCOUNT MAINTENANCE FEE" in d:
        return "Account Maintenance Fee"
    if "EBANKING SERVICE FEE" in d:
        return "E-Banking Service Fee"
    if "COT" in d:
        return "COT Charge"
    if cr > 0:
        return "Interbank Credit"
    if dr > 0:
        return "Other Debit"
    return "Other"


def categorize_omni(raw, dr=0.0, cr=0.0):
    d = raw.upper()
    if "CASH DEP BY" in d:
        return "Cash Deposit (3rd Party)"
    if "CHEQUE BOOK CHARGE" in d:
        return "Cheque Book Issuance Fee"
    if "INWARD CHQ NO" in d or "INWARD OMNIBSIC CHQ" in d or "INWARD EXPRESS CHQ" in d:
        return "Cheque Payment (Own)"
    if "SSNIT" in d:
        return "Direct Debit (SSNIT)"
    if "MONTHLY SERVICE CHARGE" in d:
        return "Monthly Service Charge"
    if "MOMO POSTRXN" in d:
        return "MoMo Settlement Credit (OmniBSIC)"
    if "ACH NRT" in d or "NRT-TRF" in d:
        return "Interbank Transfer (ACH/NRT)"
    if "CASH WD" in d or "CASH WITHDRAWAL" in d:
        return "Cash Withdrawal"
    if "PYMT" in d:
        return "Payment"
    if cr > 0:
        return "Interbank Credit"
    if dr > 0:
        return "Other Debit"
    return "Other"


def categorize_eco(raw, dr, cr):
    d = raw.upper()
    if cr < 0 or dr < 0:
        base = "Cash Deposit" if "DEP" in d or "RECEIVED" in d else "Transaction"
        return f"{base} Reversal (Negative Entry)"
    if re.search(r"DEP\s*(BY|-)", d) or "CASH DEPOSITED" in d:
        return "Cash Deposit (3rd Party)"
    if "CHEQUE DEPOSIT" in d:
        return "Cheque Deposit (Ecobank)"
    if "ECOBANK CHQ NO" in d and "RECEIVED" in d:
        return "Cheque Deposit (Ecobank)"
    if "CHEQUE WITHDRAWAL" in d:
        return "Cheque Withdrawal (Self)"
    if "GIP INCOMING" in d:
        return "GhIPSS Incoming Credit"
    if "SALARY" in d:
        return "Salary Payment"
    if "B/O" in d and cr > 0:
        return "Interbank Credit"
    if "CASH DEP" in d:
        return "Cash Deposit"
    if "CHARGE" in d or "FEE" in d or "COMM" in d:
        return "Bank Charge"
    if "LOAN" in d:
        return "Loan Transaction"
    if cr > 0:
        return "Interbank Credit"
    if dr > 0:
        return "Other Debit"
    return "Other"


def categorize_momo(type_, dr, cr):
    t = (type_ or "").upper()
    side = "Credit" if cr > 0 else "Debit"
    if t.startswith("TRANSFER"):
        return f"MoMo Transfer ({side})"
    if "PAYMENT" in t:
        return f"MoMo Payment ({side})"
    if "FLOAT" in t:
        return f"MoMo Float Transfer ({side})"
    if "PULL" in t:
        return f"MoMo Pull ({side})"
    return f"MoMo {t.title() or 'Transaction'} ({side})"


def party_eco(raw, dr, cr):
    m = re.search(r"DEP (?:BY |-)(.*?)(?: IFO| RECEIVED| REF|$)", raw, re.I)
    if m:
        return m.group(1).strip(" -")
    m = re.search(r"PAID TO (.*?)$", raw, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r"B/O (.*?)(?: IFO| IRO| RECEIVED| $|$)", raw, re.I)
    if m and cr > 0:
        return m.group(1).strip()
    return ""


def party_momo(r):
    if r["deposit_cr"] > 0:
        return (r["from_name"] or r["from_phone"]).upper()
    return (r["to_msisdn"] or r["after_name"]).upper()




def process(path):
    """Detect bank from page 1 and run the right parser. Returns dict(payload, issues, stats) or None."""
    doc = pymupdf.open(path)
    t0 = doc[0].get_text()
    base = os.path.splitext(os.path.basename(path))[0]
    if "STATEMENT\nOF\nACCOUNT" in t0 or t0.startswith("NOT FOR VISA"):
        return finish("ADB", base, doc, meta_adb(doc), ADB().parse(doc), ADB().finalize)
    if "CUSTOMER STATEMENT" in t0:
        rows = GTB().parse(doc)
        meta = meta_gt(doc)
        m = re.search(r"TOTAL:\s*([\d,]+\.\d{2})\s*([\d,]+\.\d{2})", doc[-1].get_text())
        if m:
            meta["total_debit"] = num(m.group(1))
            meta["total_credit"] = num(m.group(2))
        return finish("GTB", base, doc, meta, rows, GTB().finalize)
    if "ACCOUNT STATEMENT" in t0:
        return finish("OMNI", base, doc, meta_omni(doc), OMNI().parse(doc), OMNI().finalize)
    if "STATEMENT OF ACCOUNT" in t0:
        return finish("ECO", base, doc, meta_eco(doc), ECO().parse(doc), ECO().finalize)
    if "MTN MoMo Statement" in t0:
        rows = parse_momo(doc)
        meta = meta_momo(doc)
        meta["opening_balance"] = rows[0]["bal_before"]
        meta["closing_balance"] = rows[-1]["balance"]
        issues, stats = verify_momo(rows)
        decorate_momo(rows)
        return {"payload": build_payload("MTN Mobile Money (MoMo)", meta, rows),
                "issues": issues, "stats": stats, "base": base}
    return None


def verify(rows, opening, closing=None, tot_dr=None, tot_cr=None):
    """Resolve single-sided amounts by delta; check running balance. Returns (issues, stats)."""
    issues = []
    run = opening
    bad = 0
    resolved = 0
    for r in rows:
        d, c = r.get("withdrawal_dr") or 0.0, r.get("deposit_cr") or 0.0
        delta = round(r["balance"] - run, 2)
        if (d is None or c is None):
            side_amt = c if c is not None else d
            if delta >= 0:
                c, d = side_amt, 0.0
            else:
                d, c = side_amt, 0.0
            resolved += 1
        d, c = round(d or 0, 2), round(c or 0, 2)
        run = round(run + c - d, 2)
        if abs(run - r["balance"]) > 0.005:
            bad += 1
            if bad <= 5:
                issues.append(f"balance mismatch @{r.get('trans')}: expected {run:,.2f} got {r['balance']:,.2f} ({r.get('raw_details','')[:50]})")
        r["withdrawal_dr"], r["deposit_cr"] = d, c
    if bad:
        issues.insert(0, f"{bad}/{len(rows)} running-balance mismatches")
    td = round(sum(r["withdrawal_dr"] for r in rows), 2)
    tc = round(sum(r["deposit_cr"] for r in rows), 2)
    if tot_dr is not None and abs(td - tot_dr) > 0.01:
        issues.append(f"sumDR {td:,.2f} != stated {tot_dr:,.2f}")
    if tot_cr is not None and abs(tc - tot_cr) > 0.01:
        issues.append(f"sumCR {tc:,.2f} != stated {tot_cr:,.2f}")
    if closing is not None and abs(run - closing) > 0.01:
        issues.append(f"final balance {run:,.2f} != stated closing {closing:,.2f}")
    stats = {"rows": len(rows), "bad": bad, "resolved": resolved,
             "total_dr": td, "total_cr": tc, "final": run}
    return issues, stats


def build_payload(bank, meta, rows):
    tx = []
    for r in rows:
        tx.append({
            "trans_date": r["trans"], "ref_number": r.get("ref_number", "") or r.get("ref", "") or r.get("refid", ""),
            "raw_details": r["raw_details"], "transaction_category": r.get("cat", ""),
            "party_name": r.get("party", ""), "reference_id": r.get("refid", ""),
            "value_date": r.get("value_date", r["trans"]),
            "withdrawal_dr": round(float(r["withdrawal_dr"] or 0), 2),
            "deposit_cr": round(float(r["deposit_cr"] or 0), 2),
            "balance": round(float(r["balance"] or 0), 2),
        })
    return {"bank_name": bank, "account_summary": meta, "transactions": tx}


def finish(bank_key, base, doc, meta, rows, finalize):
    rows = finalize(rows)
    if bank_key == "ADB":
        meta["opening_balance"] = round(rows[0]["balance"] - rows[0]["deposit_cr"] + rows[0]["withdrawal_dr"], 2)
        meta["closing_balance"] = rows[-1]["balance"]
        for r in rows:
            r["cat"] = categorize_adb(r["raw_details"], r["withdrawal_dr"], r["deposit_cr"])
            r["party"] = party_adb(r["raw_details"])
            r["refid"] = r.get("ref_number", "")
    elif bank_key == "GTB":
        for r in rows:
            r["cat"] = categorize_gtb(r["raw_details"], r["withdrawal_dr"], r["deposit_cr"])
            r["party"] = party_gtb(r["raw_details"])
            r["refid"] = ""
    elif bank_key == "OMNI":
        for r in rows:
            r["cat"] = categorize_omni(r["raw_details"], r["withdrawal_dr"], r["deposit_cr"])
            r["party"] = party_omni(r["raw_details"])
            r["refid"] = ""
    elif bank_key == "ECO":
        run = meta["opening_balance"]
        for r in rows:
            delta = round(r["balance"] - run, 2)
            d0, c0 = r["withdrawal_dr"] or 0.0, r["deposit_cr"] or 0.0
            # keep printed values when consistent with the balance movement
            # (incl. negative credits/debits as printed for reversals)
            printed_ok = (d0 != 0 and abs(d0 + delta) < 0.005) or \
                         (c0 != 0 and abs(c0 - delta) < 0.005) or \
                         (d0 == 0 and c0 == 0 and delta == 0)
            if not printed_ok:
                if delta >= 0:
                    r["deposit_cr"], r["withdrawal_dr"] = round(delta, 2), 0.0
                else:
                    r["withdrawal_dr"], r["deposit_cr"] = round(-delta, 2), 0.0
            run = r["balance"]
        for r in rows:
            r["cat"] = categorize_eco(r["raw_details"], r["withdrawal_dr"], r["deposit_cr"])
            r["party"] = party_eco(r["raw_details"], r["withdrawal_dr"], r["deposit_cr"])
            m = ECO_REF.search(r["raw_details"])
            r["refid"] = m.group(1) if m else ""
    issues, stats = verify(rows, meta["opening_balance"], meta.get("closing_balance"),
                           meta.get("total_debit"), meta.get("total_credit"))
    payload = build_payload({"ADB": "ADB Bank (Agricultural Development Bank)",
                             "GTB": "Guaranty Trust Bank (Ghana) Ltd",
                             "OMNI": "OmniBSIC Bank Ghana Ltd",
                             "ECO": "Ecobank Ghana PLC"}[bank_key], meta, rows)
    return {"payload": payload, "issues": issues, "stats": stats, "base": base}


def verify_momo(rows):
    issues, bad, dm = [], 0, 0
    prev_after = None
    for i, r in enumerate(rows):
        if prev_after is not None and abs(r["bal_before"] - prev_after) > 0.005:
            bad += 1
            if bad <= 5:
                issues.append(f"continuity break row {i}: prev after {prev_after:,.2f} vs bal before {r['bal_before']:,.2f}")
        if not r["delta_match"]:
            dm += 1
        prev_after = r["balance"]
    if dm:
        issues.append(f"{dm} rows where |balance delta| != amount (fee/reversal rows)")
    td = round(sum(r["withdrawal_dr"] for r in rows), 2)
    tc = round(sum(r["deposit_cr"] for r in rows), 2)
    return issues, {"rows": len(rows), "bad": bad, "resolved": dm,
                    "total_dr": td, "total_cr": tc, "final": rows[-1]["balance"]}


def decorate_momo(rows):
    for r in rows:
        r["cat"] = categorize_momo(r["type"], r["withdrawal_dr"], r["deposit_cr"])
        r["party"] = party_momo(r)
        r["refid"] = r["ref"]
        r["ref_number"] = r["ref"]


# party extractors
def _after(raw, key, stop=None):
    i = raw.upper().find(key)
    if i < 0:
        return ""
    rest = raw[i + len(key):]
    if stop:
        for s in stop:
            j = rest.upper().find(s)
            if j >= 0:
                rest = rest[:j]
    return rest.strip(" ,.-")


def party_gtb(raw):
    bo = _after(raw, "B/O ", [" The zone", "The zone", "Remarks"])
    if bo:
        bo = re.sub(r"\s+", " ", bo).upper()[:60]
        half = len(bo) // 2
        if bo[:half] == bo[half:]:
            bo = bo[:half]
        return bo
    return ""


def party_adb(raw):
    d = raw.upper()
    if "IFO " in d:
        return _after(raw, "IFO ").upper().replace(" FOR", "", 1).strip()[:60]
    if "B/O " in d:
        return _after(raw, "B/O ").upper()[:60]
    if "FROM " in d:
        return _after(raw, "FROM ").upper()[:60]
    if "MTN-PULL" in d:
        m = re.search(r"MTN-PULL\s+(\d+)", raw)
        return f"MTN MOMO {m.group(1)}" if m else "MTN MOMO"
    return ""


def party_omni(raw):
    d = raw.upper()
    if " BY " in d:
        return _after(raw, " BY ", [" AT "]).upper()[:60]
    if "B/O" in d:
        return _after(raw, "B/O ").upper()[:60]
    if "MOMO POSTRXN" in d:
        return "MTN MOMO (SETTLEMENT)"
    return ""


# ============================ firstbank parser ========================
"""FirstBank Ghana statement parser (ported 1:1 from the verified pipeline)."""
import os
import re
import tempfile

try:
    import pymupdf
except Exception as _e:  # guarded: reported via /api/extract health
    pymupdf = None
    _IMPORT_ERRS.append(f"pymupdf: {_e}")


import re
import json
import sys
try:
    import pymupdf
except Exception as _e:  # guarded: reported via /api/extract health
    pymupdf = None
    _IMPORT_ERRS.append(f"pymupdf: {_e}")

DATE_RE = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")

def parse(pdf_path=None):
    doc = pymupdf.open(pdf_path or PDF)
    all_rows = []
    opening = closing = None
    for pno, page in enumerate(doc, start=1):
        words = page.get_text("words")  # x0,y0,x1,y1,text,block,line,wno
        # drop diagonal-watermark fragments at fixed positions on every page
        words = [w for w in words if not (
            (w[4] == "FirstBank" and abs(w[1] - 355.99) < 1.5) or
            (w[4] == "Ghana" and abs(w[1] - 173.60) < 1.5))]

        # group words into visual lines by y
        lines = {}
        for w in words:
            key = round(w[1] / 3.0)  # 3pt bucket; line pitch is 6pt, rows 20pt
            lines.setdefault(key, []).append(w)
        ordered = [sorted(ws, key=lambda w: w[0]) for k, ws in sorted(lines.items())]

        # locate header line to derive column edges
        edges = None
        header_seen = False
        for ws in ordered:
            texts = [w[4] for w in ws]
            if "Trans" in texts and any(t.startswith("Withdrawal") for t in texts):
                def x(tok_prefix):
                    return next(w[0] for w in ws if w[4].startswith(tok_prefix))
                edges = [x("Trans"), x("Ref."), x("Transaction"), x("Value"),
                         x("Withdrawal"), x("Deposit"), x("Balance")]
                header_seen = True
                continue
            if not header_seen or not edges:
                continue  # bank name / blank lines above header
            joined = " ".join(texts)
            if re.match(r"^Page \d+ of \d+$", joined) or joined.strip() == "FirstBank Ghana":
                continue

            # bucket words into columns
            cols = [[] for _ in range(7)]
            for w in ws:
                x0 = w[0]
                idx = 6
                for i in range(6):
                    if x0 < edges[i + 1] - 2:
                        idx = i
                        break
                # right-edge correction: DR/CR/Balance numbers are right-aligned;
                # a word starting before its column header belongs to previous numeric col
                cols[idx].append(w)
            trans, refn, det, val, dr, cr, bal = [" ".join(w[4] for w in c) for c in cols]

            if trans in ("Opening Balance", "") and det in ("Opening Balance", "Closing Balance") :
                amt = float(bal.replace(",", "")) if bal else None
                if det == "Opening Balance":
                    opening = amt
                else:
                    closing = amt
                continue
            if not DATE_RE.match(trans):
                if det and all_rows and not refn and not val and not dr and not cr and not bal:
                    all_rows[-1]["det_lines"].append(det)  # wrapped description line
                elif det or refn:
                    print(f"  [warn p{pno}] unparsed line: T='{trans}' R='{refn}' D='{det}' V='{val}' DR='{dr}' CR='{cr}' B='{bal}'")
                continue

            all_rows.append({
                "page": pno, "trans": trans, "refn": refn.strip(),
                "det_lines": [det] if det else [],
                "value": val if DATE_RE.match(val) else (trans if not val else val),
                "dr": dr, "cr": cr, "bal": bal,
            })

    # merge wrapped description lines
    for r in all_rows:
        merged = ""
        for ln in r["det_lines"]:
            ln = ln.strip()
            if not ln:
                continue
            if merged.endswith("-"):
                merged += ln
            else:
                merged += (" " if merged else "") + ln
        r["details"] = re.sub(r"\s+", " ", merged).strip()
        del r["det_lines"]
    return all_rows, opening, closing

def num(s):
    s = (s or "0.00").replace(",", "").strip()
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    return float(s) if s else 0.0


DATE_RE = re.compile(r"\d{2}-[A-Za-z]{3}-\d{4}")
PERIOD_RE = re.compile(r"statement for the period:\s*(\d{2}-[A-Za-z]{3}-\d{4})\s*To\s*(\d{2}-[A-Za-z]{3}-\d{4})")
AMT_RE = re.compile(r"^\(?([\d,]+\.\d{2})\)?$")
ADDR_STOP = "please find below"

LABELS = ["Account No", "Account Name", "Account Type", "Currency",
          "Available Balance", "Total Credit", "Total Debit", "Pending Debit"]

AMT_RE = re.compile(r"^\(?([\d,]+\.\d{2})\)?$")
ADDR_STOP = "please find below"


def parse_amount(s):
    s = (s or "").strip()
    m = AMT_RE.match(s)
    if not m:
        return None
    v = float(m.group(1).replace(",", ""))
    return -v if s.startswith("(") else v


def extract_header(page_text):
    """Parse the letter-header block of a statement section into metadata."""
    lines = [l.strip() for l in page_text.splitlines() if l.strip()]
    meta = {"bank_name": lines[0] if lines else ""}
    meta["account_holder_greeting"] = ""
    addr_lines, in_addr = [], False
    i = 0
    while i < len(lines):
        ln = lines[i]
        if ln.startswith("Dear "):
            meta["account_holder_greeting"] = ln[5:].strip()
            in_addr = True
        elif in_addr and ln.lower().startswith(ADDR_STOP):
            in_addr = False
            m = PERIOD_RE.search(ln)
            if m:
                meta["statement_period_start"], meta["statement_period_end"] = m.groups()
        elif in_addr:
            addr_lines.append(ln)
        elif ln.rstrip(":") in LABELS and i + 1 < len(lines):
            meta[ln.rstrip(":")] = lines[i + 1].strip()
            i += 1
        i += 1

    def clean_address(raw_lines):
        text = re.sub(r"\s+", " ", " ".join(raw_lines)).strip(" ,")
        toks = [t.strip() for t in text.split(",") if t.strip()]
        # collapse duplicated words inside a token ("Accra Accra" -> "Accra")
        collapsed = []
        for t in toks:
            parts, ded = t.split(), []
            for w in parts:
                if not ded or w.upper() != ded[-1].upper():
                    ded.append(w)
            collapsed.append(" ".join(ded))
        # dedupe repeats, keep numbers, drop bare fragments contained in later tokens
        seen, out = set(), []
        for t in collapsed:
            key = re.sub(r"[^A-Z0-9]", "", t.upper())
            if not key or (key in seen and not key.isdigit()):
                continue
            seen.add(key)
            out.append(t)
        out = [t for i, t in enumerate(out)
               if not any(o.lower().startswith(t.lower() + " ") for o in out[i+1:])]
        titled = []
        for t in out:
            words = t.split()
            titled.append(" ".join(w if (i == 0 and w.lower() == "no") else w.title()
                                   for i, w in enumerate(words)))
        addr = ", ".join(titled)
        return re.sub(r"^No\s+A\.249/\s*,?\s*", "No A.249/, ", addr)

    summary = {
        "account_name": meta.get("Account Name", meta.get("account_holder_greeting", "")),
        "customer_address": clean_address(addr_lines),
        "account_number": meta.get("Account No", ""),
        "account_type": meta.get("Account Type", ""),
        "currency": meta.get("Currency", ""),
        "statement_period_start": meta.get("statement_period_start", ""),
        "statement_period_end": meta.get("statement_period_end", ""),
        "opening_balance": None,   # filled from the table's Opening Balance row
        "closing_balance": None,   # filled from the table's Closing Balance row
        "available_balance": parse_amount(meta.get("Available Balance", "")) or 0.0,
        "total_credit": parse_amount(meta.get("Total Credit", "")) or 0.0,
        "total_debit": parse_amount(meta.get("Total Debit", "")) or 0.0,
    }
    return meta["bank_name"], summary





def find_sections(doc):
    """Return list of (start_page, end_page_inclusive) - one per letter header found."""
    starts = []
    for i, page in enumerate(doc):
        if "Please find below your bank statement for the period" in page.get_text():
            starts.append(i)
    if not starts:
        return [(0, len(doc) - 1)]
    sections = []
    for n, s in enumerate(starts):
        e = (starts[n + 1] - 1) if n + 1 < len(starts) else len(doc) - 1
        sections.append((s, e))
    return sections





REF_TOK = re.compile(r"Ref\S+")

def split_ref(details):
    m = REF_TOK.search(details)
    return m.group(0) if m else ""

def after_colon(details, key):
    """Text after 'KEY:' up to the Ref token, e.g. CASH W/D 3RD PARTY:JOHN X Ref12 -> 'JOHN X'."""
    i = details.find(key)
    if i < 0:
        return ""
    rest = details[i + len(key):]
    rest = re.split(r"\s+Ref\S+$", rest)[0]
    return rest.strip(" .")

def categorize(details, dr=0.0, cr=0.0):
    d = details.lstrip(":").strip()
    dl = d.upper()
    # --- Cheque deposits from other banks (generic) ---
    m = re.match(r"^([A-Z][A-Z ]*?)\s*CHQ\s*#\s*\d+\s*B\s*/?\s*O\s+(.*?)\s+Ref\S*$", d)
    if m:
        bank = " ".join(m.group(1).split())
        bank = {"ECO": "Ecobank", "ECOBANK": "Ecobank", "STAN CHAT": "Stanbic", "STAN": "Stanbic",
                "UMB": "UMB", "ABSA": "Absa", "NIB": "NIB", "UBA": "UBA", "ADB": "ADB", "BOA": "BOA"}.get(bank, bank.title())
        return f"Cheque Deposit ({bank})", m.group(2).strip()
    # --- Cash ---
    if dl.startswith("CASH W/D 3RD PARTY:"):
        return "Cash Withdrawal (3rd Party)", after_colon(d, "PARTY:")
    if dl.startswith("CASH W/D SELF:"):
        return "Cash Withdrawal (Self)", after_colon(d, "SELF:")
    if dl.startswith("CASH DEP 3RD PARTY:"):
        return "Cash Deposit (3rd Party)", after_colon(d, "PARTY:")
    if dl.startswith("CASH DEPOSIT:"):
        return "Cash Deposit", after_colon(d, "DEPOSIT:")
    # --- Transfers ---
    if "FUNDS TRSF:IFO" in dl:
        return "Funds Transfer (In Favour Of)", after_colon(d, "IFO ").split(" Ref")[0].strip() or after_colon(d, "IFO")
    if "FUNDS TRSF:B/O" in dl:
        return "Funds Transfer (On Behalf Of)", after_colon(d, "B/O ")
    if dl.startswith("NRT:TRF"):
        return "Interbank Transfer (NRT)", "AVE MARIA SCHOOL"
    # --- Cheque unpaid cycle ---
    if dl.startswith("COMM ON UNPAID"):
        return "Unpaid Cheque Commission", ""
    if dl.startswith("UNPAID/CHQ"):
        return "Unpaid Cheque Return (Refer to Drawer)", ""
    if "RTD CHARGE" in dl:
        return "Cheque Return Charge (RTD 10%)", ""
    if dl.startswith("INW CLEARING CHQ"):
        return "Inward Clearing (Own Cheque Paid)", ""
    if dl.startswith("INW CLG"):
        return "Inward Clearing (Cheque Presentment)", ""
    if dl.startswith("INW CLEARING"):
        return "Inward Clearing", ""
    # --- fees & charges (recurring) ---
    if dl.startswith("SMS ALERT CHARGE"):
        return "SMS Alert Charge", ""
    if dl.startswith("TOKEN MAINTAINANCE") or dl.startswith("TOKEN MAINTENANCE"):
        return "Token Maintenance Charge", ""
    if dl.startswith("SBS:STATEMENTCHARGES"):
        return "Statement Charge (SBS)", ""
    # --- NRT interbank bulk items ---
    if dl.startswith("NRT:CHG:"):
        return "Interbank Transfer Charge (NRT)", ""
    if dl.startswith("NRT:NRT BULK"):
        return "Interbank Bulk Debit (NRT)", ""
    # --- direct debits as advised ---
    if dl.startswith("DR AS ADVISED:"):
        who = after_colon(d, "ADVISED:")
        party = who.split()[0] if who.split() else ""
        return "Direct Debit (As Advised)", party
    # --- outward cheque return ---
    if dl.startswith("OUTRET/"):
        return "Outward Cheque Return (Refer to Drawer)", ""
    # --- cheque deposit adjustments ---
    if dl.startswith("CHQ DEP:IFO"):
        return "Cheque Deposit Adjustment", after_colon(d, "IFO ")
    # --- squashed school-fee credits ---
    m = re.match(r"^0+([A-Za-z ]*fees)\s+Ref\S*$", d, re.I)
    if m:
        return "Interbank Credit (School Fees)", ""
    if dl.startswith("FBM:"):
        return "Mobile Transfer Credit (FBM)", after_colon(d, "FBM:")
    # --- Fees ---
    if dl.startswith("E-BUNDLE CHARGE"):
        return "E-Bundle Charge", ""
    if "CHEQUE BOOK ISSUANCE" in dl:
        return "Cheque Book Issuance Fee", ""
    if dl.startswith("MONTHLY SALARY"):
        return "Salary Payment (FTM)", ""
    # --- GHIPSS ---
    if "GHIPSS:CHG:" in dl:
        return "GHIPSS Channel Charge", after_colon(d, "Channels:").split("/")[0].strip()
    if "GHIPSS:OTHER_CHANNELS:" in dl:
        return "GHIPSS Transfer (Other Channels)", after_colon(d, "Channels:").split("/")[0].strip()
    m = re.match(r"^P03090", d)
    if m:
        return "GhIPSS Interbank Credit (GIP)", ""
    m = re.match(r"^(P03062)Outward transfer/(?:\S+?)/Fee", d)
    if m:
        return "GhIPSS Outward Transfer (P03062)", ""
    m = re.match(r"^(P03040)(.+?)\s+\S*emerald", d)
    if m:
        return f"GhIPSS Direct Credit ({m.group(1)})", m.group(2).strip()
    m = re.match(r"^(P03\d+)FTR-(.+?)(?:\s+Ref\S*)?$", d)
    if m:
        party = re.split(r"\s+FEES-\d+$", m.group(2).strip())[0].strip()
        return f"GhIPSS Direct Credit ({m.group(1)})", party
    m = re.match(r"^(P03[0-9]{3})", d)
    if m:
        return f"GhIPSS Direct Credit ({m.group(1)})", ""
    # --- Transfers (generic) ---
    if dl.startswith("FUNDS TRSF:"):
        return "Funds Transfer", after_colon(d, "TRSF:").split(" Ref")[0].strip()
    # --- Squashed mobile/interbank names ---
    if "JEREMIAH" in dl.upper():
        return "Interbank Credit", "JEREMIAH VROOM"
    if "THEOPHILUS" in dl:
        return "Interbank Credit", "THEOPHILUS AND FELICIA AM"
    # --- direction-aware fallback ---
    name = REF_TOK.sub("", d).strip(" :/-0123456789").strip()
    if cr > 0 and name:
        return "Interbank Credit", name
    if dr > 0 and name:
        return "Other Debit", name
    return "Other", ""



def parse_page_range(doc, s, e):
    """parse() over doc pages [s..e] via a temp single-section pdf."""
    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    tmp.close()
    nd = pymupdf.open()
    nd.insert_pdf(doc, from_page=s, to_page=e)
    nd.save(tmp.name)
    nd.close()
    try:
        rows, opening, closing = parse(tmp.name)
    finally:
        os.unlink(tmp.name)
    return rows, opening, closing


def extract_fbn(pdf_path):
    """Full FBN extraction -> (payload, issues, stats). Handles the first section found."""
    doc = pymupdf.open(pdf_path)
    s, e = find_sections(doc)[0]
    bank, summary = extract_header(doc[s].get_text())
    rows_r, op, cl = parse_page_range(doc, s, e)
    if op is None or cl is None:
        return None, ["opening/closing balance row not found"], {}
    run, bad, tdr, tcr = op, 0, 0.0, 0.0
    for r in rows_r:
        d, c, b = num(r["dr"]), num(r["cr"]), num(r["bal"])
        tdr += d; tcr += c
        run = round(run + c - d, 2)
        if abs(run - b) > 0.005:
            bad += 1
    issues = []
    if bad:
        issues.append(f"{bad}/{len(rows_r)} running-balance mismatches")
    summary["opening_balance"] = op
    summary["closing_balance"] = cl
    if not summary["total_debit"]:
        summary["total_debit"] = round(tdr, 2)
    if not summary["total_credit"]:
        summary["total_credit"] = round(tcr, 2)
    if abs(tdr - summary["total_debit"]) > 0.01:
        issues.append(f"sumDR {tdr:,.2f} != header {summary['total_debit']:,.2f}")
    if abs(tcr - summary["total_credit"]) > 0.01:
        issues.append(f"sumCR {tcr:,.2f} != header {summary['total_credit']:,.2f}")
    transactions = []
    for r in rows_r:
        details = r["details"]
        cat, party = categorize(details, num(r["dr"]), num(r["cr"]))
        party = re.sub(r"\s+", " ", party).upper().strip() if party else ""
        transactions.append({
            "trans_date": r["trans"], "ref_number": r["refn"], "raw_details": details,
            "transaction_category": cat, "party_name": party,
            "reference_id": split_ref(details), "value_date": r["value"],
            "withdrawal_dr": round(num(r["dr"]), 2), "deposit_cr": round(num(r["cr"]), 2),
            "balance": round(num(r["bal"]), 2),
        })
    payload = {"bank_name": bank, "account_summary": summary, "transactions": transactions}
    stats = {"rows": len(transactions), "bad": bad, "total_dr": round(tdr, 2),
             "total_cr": round(tcr, 2), "final": cl}
    return payload, issues, stats


# ============================ format dispatcher =======================
"""Sniff bank format from page 1 and dispatch to the right parser."""
import os
import tempfile

try:
    import pymupdf
except Exception as _e:  # guarded: reported via /api/extract health
    pymupdf = None
    _IMPORT_ERRS.append(f"pymupdf: {_e}")



def extract_any(pdf_bytes: bytes, filename: str = "statement.pdf"):
    """Returns dict: {bank_key, payload, issues, stats, pages} or raises ValueError."""
    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    try:
        tmp.write(pdf_bytes)
        tmp.close()
        doc = pymupdf.open(tmp.name)
        if not any(p.get_text().strip() for p in doc):
            raise ValueError("Scanned/image-only PDF (no text layer). OCR is not supported yet.")
        pages = len(doc)
        doc.close()

        t0 = pymupdf.open(tmp.name)[0].get_text()
        if "Please find below your bank statement" in t0 or (
                "FirstBank" in t0 and "Dear " in t0):
            payload, issues, stats = extract_fbn(tmp.name)
            bank_key = "FBN"
        else:
            r = process(tmp.name)
            if r is None:
                raise ValueError(
                    "Unrecognised statement format. Supported: FBN, GTBank, ADB, OmniBSIC, "
                    "Ecobank, MTN MoMo. Send a sample and the parser can be extended.")
            payload, issues, stats = r["payload"], r["issues"], r["stats"]
            bank_key = {"ADB": "ADB", "GTB": "GTB", "OMNI": "OMNI", "ECO": "ECO"}[
                r["base"][:3].upper()] if False else _bank_key_from_payload(payload)
        if payload is None:
            raise ValueError("Could not find statement table in this PDF.")
        return {"bank_key": bank_key, "payload": payload, "issues": issues,
                "stats": stats, "pages": pages}
    finally:
        os.unlink(tmp.name)


def _bank_key_from_payload(payload):
    name = (payload.get("bank_name") or "").upper()
    if "ADB" in name or "AGRICULTURAL" in name:
        return "ADB"
    if "GUARANTY" in name or "GTB" in name:
        return "GTB"
    if "OMNIBSIC" in name:
        return "OMNI"
    if "ECOBANK" in name:
        return "ECO"
    if "MTN" in name or "MOMO" in name:
        return "MOMO"
    return "OTHER"


# ============================ persistence pipeline ====================
"""Persist an extracted statement into Supabase (storage + Postgres) + build XLSX."""
import hashlib
import re
from datetime import datetime



def dk(s):
    """DD-Mon-YYYY -> YYYY-MM-DD (or None)."""
    if not s:
        return None
    for fmt in ("%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def txn_hash(org_id, account_number, bank, t):
    basis = "|".join([
        str(org_id), str(account_number), str(bank),
        str(t.get("trans_date", "")), str(t.get("raw_details", "")),
        f"{t.get('withdrawal_dr') or 0:.2f}", f"{t.get('deposit_cr') or 0:.2f}",
        f"{t.get('balance') or 0:.2f}",
    ])
    return hashlib.sha256(basis.encode()).hexdigest()[:32]


def persist(sb, org_id, user_id, pdf_bytes, filename, extracted):
    """extracted: dispatcher result. Returns dict for the API response."""
    payload = extracted["payload"]
    s = payload["account_summary"]
    tx = payload["transactions"]
    bank = payload["bank_name"]
    acct_no = s.get("account_number", "")
    period_start, period_end = dk(s.get("statement_period_start")), dk(s.get("statement_period_end"))

    file_sha = hashlib.sha256(pdf_bytes).hexdigest()
    dedupe_key = f"{acct_no}|{period_start}|{period_end}|{file_sha[:16]}"

    # 1) duplicate import check
    existing = sb.select("statements", f"select=id,original_filename&dedupe_key=eq.{dedupe_key}&limit=1")
    if existing:
        return {"status": "DUPLICATE", "statement_id": existing[0]["id"],
                "message": f"Already imported as '{existing[0].get('original_filename')}'."}

    # 2) upsert account
    acc_rows = sb.insert("accounts", [{
        "org_id": org_id, "bank_name": bank, "account_number": acct_no,
        "account_name": s.get("account_name", ""), "account_type": s.get("account_type", ""),
        "currency": s.get("currency", "GHS"),
    }], upsert=True)
    account_id = acc_rows[0]["id"] if acc_rows else None

    # 3) store original pdf
    pdf_path = f"{org_id}/{file_sha}_{re.sub(r'[^A-Za-z0-9._-]+', '_', filename)}"
    sb.storage_upload("statements", pdf_path, pdf_bytes, "application/pdf")

    # 4) statement row
    warnings = extracted["issues"]
    st_rows = sb.insert("statements", [{
        "org_id": org_id, "account_id": account_id, "bank_name": bank,
        "account_number": acct_no, "account_name": s.get("account_name", ""),
        "currency": s.get("currency", "GHS"),
        "period_start": period_start, "period_end": period_end,
        "opening_balance": s.get("opening_balance"), "closing_balance": s.get("closing_balance"),
        "available_balance": s.get("available_balance"),
        "total_debit": s.get("total_debit"), "total_credit": s.get("total_credit"),
        "row_count": len(tx), "pages": extracted.get("pages"),
        "original_filename": filename, "file_sha256": file_sha,
        "storage_path": pdf_path, "dedupe_key": dedupe_key,
        "status": "OK_WITH_WARNINGS" if warnings else "OK",
        "verification": {"issues": warnings, "stats": extracted["stats"]},
        "created_by": user_id,
    }])
    statement_id = st_rows[0]["id"]

    # 5) transactions
    rows = []
    for i, t in enumerate(tx):
        rows.append({
            "statement_id": statement_id, "org_id": org_id, "account_id": account_id,
            "row_index": i,
            "trans_date": dk(t.get("trans_date")) or period_start,
            "value_date": dk(t.get("value_date")) or dk(t.get("trans_date")) or period_start,
            "ref_number": t.get("ref_number", ""), "raw_details": t.get("raw_details", ""),
            "category": t.get("transaction_category", ""), "party": t.get("party_name", ""),
            "reference_id": t.get("reference_id", ""),
            "withdrawal": t.get("withdrawal_dr") or 0, "deposit": t.get("deposit_cr") or 0,
            "balance": t.get("balance") or 0,
            "txn_hash": txn_hash(org_id, acct_no, bank, t),
        })
    for i in range(0, len(rows), 500):
        sb.insert("transactions", rows[i:i + 500], returning="return=minimal")

    # 6) xlsx export -> storage
    xls = build_xlsx(payload)
    xlsx_path = f"{org_id}/{statement_id}.xlsx"
    sb.storage_upload("exports", xlsx_path, xls,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # 7) extraction log
    sb.insert("extraction_logs", [{
        "statement_id": statement_id, "org_id": org_id,
        "bank_key": extracted.get("bank_key"), "warnings": warnings,
        "stats": extracted["stats"],
    }], returning="return=minimal")

    return {"status": "OK", "statement_id": statement_id, "account_id": account_id,
            "rows": len(tx), "warnings": warnings,
            "summary": {"bank": bank, "account_number": acct_no,
                        "period": f"{s.get('statement_period_start')} -> {s.get('statement_period_end')}",
                        "opening": s.get("opening_balance"), "closing": s.get("closing_balance"),
                        "total_debit": s.get("total_debit"), "total_credit": s.get("total_credit")}}


# ============================================================
# serverless handler (FastAPI)
# ============================================================
import fastapi  # noqa: E402
from fastapi import FastAPI, File, Header, HTTPException, UploadFile  # noqa: E402
from fastapi.middleware.cors import CORSMiddleware  # noqa: E402
from fastapi.responses import JSONResponse  # noqa: E402

app = FastAPI(title="Statement Intel - extract")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


def _sb():
    url = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = (os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
           or os.environ.get("NEXT_PUBLIC_SUPABASE_SERVICE_ROLE_KEY"))
    if not url or not key:
        raise HTTPException(500, "Server misconfigured: SUPABASE_URL / service key env vars missing. "
                                 "Check the Vercel <-> Supabase integration or add them manually.")
    return SB(url, key)


def _auth(sb, authorization):
    if os.environ.get("DEV_BYPASS_AUTH") == "1":
        return "dev-user", "dev-org"
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "Missing bearer token - sign in first")
    user = sb.get_user(authorization.split(" ", 1)[1].strip())
    if not user:
        raise HTTPException(401, "Invalid or expired session")
    profs = sb.select("profiles", f"select=org_id&id=eq.{user['id']}&limit=1")
    if not profs:
        raise HTTPException(403, "No profile for user")
    return user["id"], profs[0]["org_id"]


@app.get("/")
@app.get("/api/extract")
async def health(id: int = 0, authorization: str = Header(None)):
    if _IMPORT_ERRS:
        return JSONResponse(status_code=500, content={
            "ok": False, "service": "statement-intel extract",
            "python": sys.version.split()[0],
            "imports": "BROKEN (see detail)", "detail": _IMPORT_ERRS[0],
            "env_present": {k: bool(os.environ.get(k)) for k in (
                "SUPABASE_URL", "NEXT_PUBLIC_SUPABASE_URL",
                "SUPABASE_SERVICE_ROLE_KEY", "NEXT_PUBLIC_SUPABASE_SERVICE_ROLE_KEY")}})

    # ---- on-demand Excel export: /api/extract?id=<statement_id> ----
    if id > 0:
        try:
            from fastapi.responses import Response
            sb = _sb()
            user_id, org_id = _auth(sb, authorization)
            srows = sb.select("statements", f"select=*&id=eq.{id}&org_id=eq.{org_id}&limit=1")
            if not srows:
                raise HTTPException(404, f"Statement {id} not found in your workspace")
            s = srows[0]
            tx = sb.select("transactions", f"select=*&statement_id=eq.{id}&order=row_index.asc")
            fnum = lambda v: float(v) if v not in (None, "") else None
            summary = {
                "account_name": s.get("account_name") or "", "customer_address": "",
                "account_number": s.get("account_number") or "", "account_type": "",
                "currency": s.get("currency") or "GHS",
                "statement_period_start": s.get("period_start") or "",
                "statement_period_end": s.get("period_end") or "",
                "opening_balance": fnum(s.get("opening_balance")),
                "closing_balance": fnum(s.get("closing_balance")),
                "available_balance": fnum(s.get("available_balance")),
                "total_credit": fnum(s.get("total_credit")),
                "total_debit": fnum(s.get("total_debit")),
            }
            payload = {"bank_name": s.get("bank_name") or "", "account_summary": summary,
                       "transactions": [{
                           "trans_date": t.get("trans_date") or "",
                           "ref_number": t.get("ref_number") or "",
                           "raw_details": t.get("raw_details") or "",
                           "transaction_category": t.get("category") or "",
                           "party_name": t.get("party") or "",
                           "reference_id": t.get("reference_id") or "",
                           "value_date": t.get("value_date") or t.get("trans_date") or "",
                           "withdrawal_dr": float(t.get("withdrawal") or 0),
                           "deposit_cr": float(t.get("deposit") or 0),
                           "balance": float(t.get("balance") or 0)} for t in tx]}
            data = build_xlsx(payload)
            base_name = re.sub(r"[^A-Za-z0-9._ -]+", "_", s.get("original_filename") or f"statement_{id}")
            fname = re.sub(r"\.pdf$", "", base_name, flags=re.I) + ".xlsx"
            return Response(content=data,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f'attachment; filename="{fname}"'})
        except HTTPException:
            raise
        except Exception:
            return JSONResponse(status_code=500, content={
                "detail": "Building the Excel file failed",
                "error": traceback.format_exc(limit=6).splitlines()[-6:]})

    return {
        "ok": True, "service": "statement-intel extract",
        "python": sys.version.split()[0], "imports": "OK", "detail": None,
        "env_present": {k: bool(os.environ.get(k)) for k in (
            "SUPABASE_URL", "NEXT_PUBLIC_SUPABASE_URL",
            "SUPABASE_SERVICE_ROLE_KEY", "NEXT_PUBLIC_SUPABASE_SERVICE_ROLE_KEY")},
    }


@app.post("/")
@app.post("/api/extract")
async def extract(file: UploadFile = File(None), authorization: str = Header(None)):
    if _IMPORT_ERRS:
        return JSONResponse(status_code=500, content={
            "detail": "Extractor libraries failed to load on the server",
            "errors": _IMPORT_ERRS})
    if file is None or not file.filename:
        raise HTTPException(400, "No file uploaded (field name must be 'file')")

    sb = _sb()
    user_id, org_id = _auth(sb, authorization)

    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file")
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(413, "File too large (max 20 MB)")
    if data[:5] != b"%PDF-":
        raise HTTPException(400, "Not a PDF file")

    try:
        extracted = extract_any(data, file.filename)
    except ValueError as e:
        raise HTTPException(422, str(e))
    except Exception:
        return JSONResponse(status_code=500, content={
            "detail": "Parser crashed while reading this PDF",
            "error": traceback.format_exc(limit=6).splitlines()[-6:]})

    try:
        result = persist(sb, org_id, user_id, data, file.filename, extracted)
    except Exception:
        return JSONResponse(status_code=500, content={
            "detail": "Parsed OK but saving to the database failed",
            "error": traceback.format_exc(limit=6).splitlines()[-6:]})
    if result.get("status") == "DUPLICATE":
        return JSONResponse(status_code=409, content=result)
    return result

