"""Schema-compliant XLSX builder (openpyxl only - no pandas)."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TX_COLS = [
    ("trans_date", "Trans Date", 12),
    ("ref_number", "Ref Number", 14),
    ("raw_details", "Original Details", 52),
    ("transaction_category", "Transaction Category", 28),
    ("party_name", "Counterparty / Person", 30),
    ("reference_id", "System Ref", 22),
    ("value_date", "Value Date", 12),
    ("withdrawal_dr", "Withdrawal (DR)", 15),
    ("deposit_cr", "Deposit (CR)", 15),
    ("balance", "Balance", 15),
]


def build_xlsx(payload) -> bytes:
    """payload: {bank_name, account_summary, transactions} -> xlsx bytes."""
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)

    ws = wb.active
    ws.title = "Account Summary"
    s = payload.get("account_summary", {})
    rows = [("Bank", payload.get("bank_name", ""))] + list(s.items())
    ws.append(["Field", "Value"])
    for k, v in rows:
        ws.append([k, v if v is not None else ""])
    for c in ws[1]:
        c.fill, c.font = hdr_fill, hdr_font
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 48
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Transactions")
    ws2.append([h for _, h, _ in TX_COLS])
    for c in ws2[1]:
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(vertical="center")
    for t in payload.get("transactions", []):
        ws2.append([t.get(k, "") for k, _, _ in TX_COLS])
    money_idx = [i + 1 for i, (k, _, _) in enumerate(TX_COLS)
                 if k in ("withdrawal_dr", "deposit_cr", "balance")]
    for i, (k, _, w) in enumerate(TX_COLS, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        if k in ("withdrawal_dr", "deposit_cr", "balance"):
            for r in range(2, ws2.max_row + 1):
                ws2.cell(row=r, column=i).number_format = "#,##0.00"
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
